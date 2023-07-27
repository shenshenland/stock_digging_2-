import openpyxl
from scipy.stats import kstest
import matplotlib.pyplot as plt
import numpy as np
def read_excel_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    column_data = []
    for i in range(2000,sheet.max_row+1):
    # for i in range(2000, sheet.max_row + 1)
    # for i in range(20, sheet.max_row + 1)
    # for i in range(293, sheet.max_row + 1)/ no matter which time you choose, still get the result that this follows the normal distribution
        print(sheet.cell(i,8).value)
        if type(sheet.cell(i,8).value) == float or sheet.cell(i,8).value == 0:
            column_data.append(sheet.cell(i,8).value)
        else:
            pass


    return column_data

def check_normality(data):
    _, p = kstest(data, 'norm')
    return p >= 0.05

def plot_histogram(data):
    plt.hist(data, bins=500, alpha=0.5, color='steelblue', edgecolor='black')
    plt.xlabel('Value')
    plt.ylabel('Frequency')
    plt.title('Histogram')
    plt.show()

def calculate_mean_and_std(data):
    mean = sum(data) / len(data)
    variance = sum((x - mean) ** 2 for x in data) / len(data)
    std_deviation = variance ** 0.5
    return mean, std_deviation

if __name__ == "__main__":
    # 输入你的Excel文件路径、工作表名和列名 /put your name of excel file
    file_path = "F:\history.xlsx"
    # sheet_name = "history"

    np.random.seed(19)

    # 生成正态分布的数据 / generate random number
    mean = 0  # 均值 /mean
    std_deviation = 1  # 标准差 /standard deviation
    sample_size = 1000  # 数据点数量 /number of sample

    data = np.random.normal(mean, std_deviation, sample_size)
    # data = read_excel_data(file_path, sheet_name)

    if check_normality(data):
        print("The data follows a normal distribution.")
    else:
        print("The data does not follow a normal distribution.")



    plt.xlabel('stock return rate')
    plt.ylabel('Frequency')
    plt.title('Tesla_history_data')
    mean, std_deviation = calculate_mean_and_std(data)
    print(f"Mean: {mean}")
    print(f"Standard Deviation: {std_deviation}")

    plt.hist(data, bins=300, alpha=0.5, color='steelblue', edgecolor='black')
    plt.show()


